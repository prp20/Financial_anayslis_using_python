# https://medium.com/@gokhalehemal11/how-to-scrape-financial-websites-using-python-requests-and-beautifulsoup-5f436318ca9

import requests
from bs4 import BeautifulSoup
import time
import csv
import glob
import openpyxl
import pandas as pd
import string

balance_sheet_keys = [None, 'Basic EPS (Rs.)', 'Diluted EPS (Rs.)', 'Cash EPS (Rs.)', 'Book Value [ExclRevalReserve]/Share (Rs.)', 'Book Value [InclRevalReserve]/Share (Rs.)', 'Dividend / Share(Rs.)', 'Revenue from Operations/Share (Rs.)', 'PBDIT/Share (Rs.)', 'PBIT/Share (Rs.)', 'PBT/Share (Rs.)', 'Net Profit/Share (Rs.)', 'PBDIT Margin (%)', 'PBIT Margin (%)', 'PBT Margin (%)', 'Net Profit Margin (%)', 'Return on Networth / Equity (%)', 'Return on Capital Employed (%)', 'Return on Assets (%)', 'Total Debt/Equity (X)', 'Asset Turnover Ratio (%)', 'Current Ratio (X)', 'Quick Ratio (X)', 'Inventory Turnover Ratio (X)', 'Dividend Payout Ratio (NP) (%)', 'Dividend Payout Ratio (CP) (%)', 'Earnings Retention Ratio (%)', 'Cash Earnings Retention Ratio (%)', 'Enterprise Value (Cr.)', 'EV/Net Operating Revenue (X)', 'EV/EBITDA (X)', 'MarketCap/Net Operating Revenue (X)', 'Retention Ratios (%)', 'Price/BV (X)', 'Price/Net Operating Revenue', 'Earnings Yield']

def get_html(url):
    """ Get the HTML of a URL """
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'lxml')  # Use lxml parser
    return soup


def create_csv(letter, company_dict):
    """ Create a CSV file for each stock price letter """
    try:
        soup = get_html(
            'https://www.moneycontrol.com/india/stockpricequote/' + letter)
        time.sleep(5)

        links = soup.find_all('a', {'class': 'bl_12'})
        for link in links:
            company_list = []
            company_list.append(link.text)
            company_list.append(link['href'])
            other_urls = scrape_quick_links(link['href'])
            for row in other_urls:
                company_list.append(row['href'])
            company_dict[company_list[0]] = company_list
            print(company_list)
        print("Success for ", letter)

    except Exception as e:
        print("Exception for ", letter, ": ", e)


def print_csv_columns():
    """ Print the contents of all CSVs """
    dataset = pd.DataFrame()
    for filename in glob.glob('stocks/*.csv'):
        data = pd.read_csv(filename, header=None, index_col=False)
        data.drop([0], axis=0, inplace=True)
        data.drop([0], axis = 1, inplace=True)
        data.dropna(inplace=True)
        dataset = pd.concat([dataset, data], ignore_index=True)
        # with open(filename, newline='') as csvfile:
        #     reader = csv.reader(csvfile, delimiter=',')
        #     for row in reader:
        #         if row[0] and row[1] is not None:
        #             print(row[0])
        #             print(row[1], row[2])
        #         else:
        #             continue
    return dataset

def scrape_table(url, stock_name, sheet_name, next_page=False):
    """ Scrape a table from a web page """
    soup = get_html(url)
    table = soup.find('table', {'class': 'mctable1'})
    rows = table.find_all('tr')

    if next_page:
        wb = openpyxl.load_workbook('{}.xlsx'.format(stock_name))
    else:
        wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheet_name

    if next_page:
        first_empty_col = sheet.max_column - 2
        for i, row in enumerate(rows):
            for j, el in enumerate(row):
                if j > 2 and j < len(row) - 3:
                    cell_ref = sheet.cell(i + 1, first_empty_col + j + 1)
                    cell_ref.value = el.string
    else:
        for row in rows:
            row_list = [el.string for el in row][:-2]
            sheet.append(row_list)

    wb.save('{}.xlsx'.format(stock_name))


def scrape_quick_links(url):
    """ Scrape quick links from a web page """
    soup = get_html(url)
    quick_links = soup.find('div', {'class': 'quick_links clearfix'})
    links = quick_links.find_all('a')
    return links

def get_company_name(url):
  soup = get_html(url)
  comp_val = soup.find('div', {'class':'inid_name'})
  return comp_val.text.split("\n")[1]
    
def get_active_href(url):
    """ Get the URL of the active page """
    soup = get_html(url)
    span_tag = soup.find('span', {'class': 'nextpaging'})
    parent_tag = span_tag.find_previous('a')
    if parent_tag:
        href = parent_tag.get('href')
        if href and href != 'javascript:void();':
            return href
    return None


def scrape_table(url, stock_name, sheet_name, next_page=False):
    """ Scrape a table from a web page """
    soup = get_html(url)
    table = soup.find('table', {'class': 'mctable1'})
    rows = table.find_all('tr')

    if next_page:
        wb = openpyxl.load_workbook('{}.xlsx'.format(stock_name))
    else:
        wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheet_name

    if next_page:
        first_empty_col = sheet.max_column - 2
        for i, row in enumerate(rows):
            for j, el in enumerate(row):
                if j > 2 and j < len(row) - 3:
                    cell_ref = sheet.cell(i + 1, first_empty_col + j + 1)
                    cell_ref.value = el.string
    else:
        for row in rows:
            row_list = [el.string for el in row][:-2]
            sheet.append(row_list)

    wb.save('{}.xlsx'.format(stock_name))

def remove_punctuations(text):
    for punctuation in string.punctuation:
        text = text.replace(punctuation, '')
    return text

def get_nifty_mc_dataset(nifty_file):
    """ Creates a dataset by combining the data from the money control web scraping and 
    nifty list from the NSE."""
    dataset = print_csv_columns()
    dataset.columns = ['Company Name', 'Company Url']
    nifty_data = pd.read_csv(nifty_file)
    nifty_data['Company Name'] =  nifty_data['Company Name'].str.replace("&", "and")
    nifty_data['company_short_form'] = nifty_data['Company Name'].str.lower().str.replace("&", "and").apply(remove_punctuations)
    dataset['Company Name'] =  dataset['Company Name'].str.replace("&", "and")
    dataset['company_short_form'] = dataset['Company Name'].str.lower().str.replace("&", "and").apply(remove_punctuations)
    new_data = dataset.merge(nifty_data,on='company_short_form' ,how='right')
    nifty_mc_dataset = new_data.drop(['Series', 'ISIN Code', 'company_short_form'], axis=1)
    return nifty_mc_dataset

def get_fundamental_data_for_company(dataset, url):
    links_set = scrape_quick_links(url)
    temp_data = {}
    for link in links:
        if link.text == "Balance Sheet" or link.text == "Profit & Loss" or link.text == "Cash Flows":
            temp_data[link.text] = link['href']

def main():
    """ Get all the stock web links from A-Z available on 
    moneycontrol and store them into 
    csv files alphabetically """

    nifty_money_control_base_dataset = get_nifty_mc_dataset('ind_nifty100list.csv')

    # """ Read a url for stock and scrape the urls of financial section
    # like Balance Sheet, Profit and Loss and, Quarterly an Yearly results,
    # Cashflow statments, etc. """
    # url = "https://www.moneycontrol.com/india/stockpricequote/refineries/relianceindustries/RI"
    # scrape_quick_links(url)

    # """ Get the financial section data which we want. In this case Balance Sheet of the stock """
    # url = "https://www.moneycontrol.com/financials/relianceindustries/balance-sheetVI/RI#RI"
    # stock_name = "RELIANCE"
    # sheet_name = "Balance Sheet"

    # """ Store all the available data of all years including searching for previous years's data if available
    # and saving them into a excel file with sheet Balance Sheet in this case """
    # scrape_table(url, stock_name, sheet_name)

    # first_entry = True
    # while url:
    #     print(url)
    #     if first_entry:
    #         scrape_table(url, stock_name, sheet_name)
    #         first_entry = False
    #     else:
    #         scrape_table(url, stock_name, sheet_name, True)
    #     url = get_active_href(url)


if __name__ == '__main__':
    main()
